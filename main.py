import os
import time
from io import BytesIO
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from selenium import webdriver
from selenium.webdriver import Keys, ActionChains
from selenium.webdriver.common.by import By
import pickle
from openpyxl.drawing.image import Image
from googletrans import Translator

translator = Translator(service_urls=['translate.googleapis.com'])


class Parser:
    def __init__(self):
        self.options = webdriver.ChromeOptions()
        self.options.add_argument('--headless')
        self.options.add_experimental_option('excludeSwitches', ['enable-automation'])
        self.options.add_experimental_option('useAutomationExtension', False)
        self.options.add_argument('--disable-blink-features=AutomationControlled')
        self.options.add_argument('--disable-dev-shm-usage')
        self.options.add_argument('--disable-lazy-loading')
        self.options.add_argument('--popup-blocking')
        self.options.add_argument('--proxy-server=')  # PROXY например: '--proxy-server=127.0.0.1'
        self.options.add_argument('--disable-gpu')
        self.options.add_argument(' "--autoplay-policy=no-user-gesture-required"')
        self.options.add_experimental_option('prefs', {
            'profile.managed_default_content_settings.popups': 2})
        self.img_counter = 1

    def get_cookies(self):
        try:
            driver = self.del_humanity_check()
            driver.get(
                'https://login.taobao.com/?redirect_url=https%3A%2F%2Flogin.1688.com%2Fmember%2Fjump.htm%3Ftarget'
                '%3Dhttps%253A%252F%252Flogin.1688.com%252Fmember%252FmarketSigninJump.htm%253FDone%253Dhttps%25253'
                'A%25252F%25252Fwww.1688.com%25252F&style=tao_custom&from=1688web')
            driver.implicitly_wait(5)
            login_box = driver.find_element(By.ID, 'fm-login-id')
            login_box.clear()
            login_box.send_keys('login')  # Логин авторизации напр.: tb9096673730
            password_box = driver.find_element(By.ID, 'fm-login-password')
            password_box.clear()
            password_box.send_keys('password')  # Пароль авторизации
            password_box.send_keys(Keys.ENTER)

            time.sleep(7)
            driver.switch_to.frame(0)

            slider = driver.find_element(By.XPATH, '//span[@class="nc_iconfont btn_slide"]')
            actions = ActionChains(driver)
            move = actions.move_to_element(slider)
            move.perform()
            time.sleep(1)
            ri = actions.drag_and_drop_by_offset(slider, 260, 0)
            ri.perform()

            time.sleep(180)
            pickle.dump(driver.get_cookies(), open('cookies', 'wb'))
            driver.close()
            driver.quit()
        except Exception as e:
            print(e)

    def del_humanity_check(self):
        driver = webdriver.Chrome(options=self.options)
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
                delete window.cdc_adoQpoasnfa76pfcZLmcfl_Promise;
                delete window.cdc_adoQpoasnfa76pfcZLmcfl_Object;
                delete window.cdc_adoQpoasnfa76pfcZLmcfl_Symbol;
                delete window.cdc_adoQpoasnfa76pfcZLmcfl_Function;
                delete window.cdc_adoQpoasnfa76pfcZLmcfl_Proxy;
                delete window.cdc_adoQpoasnfa76pfcZLmcfl_Array;
            """
        })
        return driver

    @staticmethod
    def slider_detect(driver):
        slider = driver.find_elements(By.XPATH, '//div[@class="baxia-dialog-content"]')
        if len(slider) > 0:
            frame = driver.find_elements(By.TAG_NAME, 'iframe')
            if len(frame) > 0:
                driver.switch_to.frame(0)
                slider = driver.find_element(By.XPATH, '//span[@class="nc_iconfont btn_slide"]')
                actions = ActionChains(driver)
                move = actions.move_to_element(slider)
                move.perform()
                time.sleep(1)
                ri = actions.drag_and_drop_by_offset(slider, 260, 0)
                ri.perform()
                time.sleep(5)
                driver.refresh()

    def parse(self):
        driver = self.del_humanity_check()
        driver.maximize_window()
        driver.get('https://1688.com/')
        driver.implicitly_wait(5)
        self.slider_detect(driver)
        for cookie in pickle.load(open('cookies', 'rb')):
            driver.add_cookie(cookie)
        time.sleep(2)
        driver.refresh()
        driver.implicitly_wait(5)
        self.slider_detect(driver)
        self.get_by_img(driver, 'Зубная щетка',
                                'АБСОЛЮТНЫЙ ПУТЬ ДО КАРТИНКИ')  # Поиск по картинке
        self.get_by_text(driver, 'Зубная щетка')  # Тут ваш поисковый запрос на русском

        driver.close()
        driver.quit()

    def get_by_img(self, driver, product, path):
        self.create_excel(product.replace(' ', '-'))
        file_input = driver.find_element(By.XPATH, '//div[@class="react-file-reader"]').find_element(
            By.XPATH, "//input[@type='file']")
        print(file_input.get_attribute('id'))
        file_input.send_keys(f'{path}')
        actions = ActionChains(driver)
        for k in range(15):  # количество страниц которые нужно пролистать
            print('k:', k)
            actions.key_down(Keys.END).key_up(Keys.END).perform()
            self.slider_detect(driver)
            time.sleep(1)
            driver.implicitly_wait(3)
        self.slider_detect(driver)
        driver.implicitly_wait(10)
        self._get_offers_links(driver)

    def get_by_text(self, driver, text):
        self.create_excel(text.replace(' ', '-'))
        product_links = []
        search_box = driver.find_element(By.ID, 'home-header-searchbox')
        search_box.clear()
        text = translator.translate(text=text, dest='zh-cn').text
        search_box.send_keys(text)
        search_box.send_keys(Keys.ENTER)
        driver.implicitly_wait(5)
        self.slider_detect(driver)
        driver.close()
        driver.switch_to.window(driver.window_handles[1])
        driver.implicitly_wait(5)
        # time.sleep(555)
        actions = ActionChains(driver)
        self.slider_detect(driver)
        for i in range(15):  # Количество страниц которые нужно пролистать
            for k in range(3):
                actions.key_down(Keys.END).key_up(Keys.END).perform()
                self.slider_detect(driver)
                time.sleep(2)
                driver.implicitly_wait(3)

            response = driver.page_source
            soup = BeautifulSoup(response, 'lxml')
            offer_list = soup.find('ul', class_="offer-list").find_all(
                'div', class_='mojar-element-title')

            for product in offer_list:
                link = product.find('a').get('href')
                if 'detail' not in link:
                    continue
                product_links.append(link)

            driver.find_element(By.XPATH, '//span[@class="fui-paging-list"]').find_element(
                By.XPATH, '//a[@class="fui-next"]').click()

            time.sleep(2)

            self.slider_detect(driver)
        self.get_detail_info(driver, product_links)

    def get_detail_info(self, driver, links):
        for link in links:
            driver.get(link)
            self.slider_detect(driver)
            self.mini_login_page(driver)
            driver.implicitly_wait(5)
            response = driver.page_source
            soup = BeautifulSoup(response, 'lxml')
            print(link)
            title = soup.find_all('div', attrs={'class': 'title-content'})
            if len(title) > 0:
                title = title[0].find(
                    'div', attrs={'class': 'title-text'}).text.strip()
            else:
                continue
            title = translator.translate(title, dest='ru').text
            pictures = soup.find_all(
                'img', attrs={'class': 'detail-gallery-img'})
            pictures = [img.get('src') for img in pictures if 'tbvideo' not in img.get('src')]

            price = soup.find_all(
                'span', attrs={'class': 'price-text'})
            if len(price) > 0:
                price = price[0].text.strip()
            else:
                continue
            url = link

            data = [['', title, price, ", ".join(pictures), url]]
            self.add_data_to_excel(data, 'Зубная-щетка', pictures)
        driver.close()

    def add_data_to_excel(self, data, name, pictures):
        workbook = load_workbook(f'./docs/{name}.xlsx')
        sheet = workbook.active
        response = requests.get(pictures[1])

        try:
            image_content = response.content
            img = Image(BytesIO(image_content))

            img.width = 100
            img.height = 100

            for row in data:
                sheet.append(row)
            cell = sheet.cell(row=self.img_counter + 1, column=1)
            sheet.row_dimensions[self.img_counter + 1].height = 77
            sheet.add_image(img, cell.coordinate)
            self.img_counter += 1

            workbook.save(f'./docs/{name}.xlsx')
        except Exception:
            return None

    @staticmethod
    def create_excel(title):
        path = os.path.join(os.path.dirname(__file__), f'./docs/{title}.xlsx')
        if not os.path.exists(path):
            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'Превью'
            ws['B1'] = 'Наименование товара'
            ws['C1'] = 'Цена'
            ws['D1'] = 'Изображения'
            ws['E1'] = 'Ссылка на товар'

            ws.column_dimensions['A'].height = 17
            ws.column_dimensions['B'].height = 100
            ws.column_dimensions['C'].height = 7
            ws.column_dimensions['D'].height = 35
            ws.column_dimensions['E'].height = 42

            wb.save(f'./docs/{title}.xlsx')

    def _get_offers_links(self, driver):
        product_links = []
        response = driver.page_source
        soup = BeautifulSoup(response, 'lxml')
        offer_list = soup.find('ul', class_="offer-list").find_all(
            'div', class_='mojar-element-title')

        for product in offer_list:
            link = product.find('a').get('href')
            if 'detail' not in link:
                continue
            product_links.append(link)

        self.slider_detect(driver)
        self.get_detail_info(driver, product_links)

    @staticmethod
    def mini_login_page(driver):
        login_page = driver.find_elements(By.XPATH, '//body[@class=mini-login-body]')
        if len(login_page) > 0:
            driver.find_element(By.XPATH, '//a[@class=show-pwd-login-link]').click()
            driver.implicitly_wait(3)
            login_box = driver.find_element(By.ID, 'fm-login-id')
            login_box.clear()
            login_box.send_keys('login')
            password_box = driver.find_element(By.ID, 'fm-login-password')
            password_box.clear()
            password_box.send_keys('password')
            driver.find_element(By.XPATH, '//input[@id=fm-agreement-checkbox]').click()
            time.sleep(1)
            driver.find_element(By.TAG_NAME, 'button').click()


parser = Parser()
parser.parse()
